from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import os
from typing import List, Dict
import mimetypes

# Document processing libraries
import PyPDF2
from docx import Document
import openpyxl
import csv

class DriveHandler:
    """Handles Google Drive API interactions with service account"""
    
    def __init__(self):
        """Initialize Google Drive API with service account credentials"""
        import json
        
        # Try to load from JSON string in environment variable first
        service_account_json = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')
        service_account_file = os.getenv('GOOGLE_SERVICE_ACCOUNT_FILE')
        
        # Define the required scopes
        SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        
        if service_account_json:
            # Load from JSON string
            try:
                cred_info = json.loads(service_account_json)
                credentials = service_account.Credentials.from_service_account_info(
                    cred_info, scopes=SCOPES
                )
            except json.JSONDecodeError:
                raise ValueError("Invalid JSON in GOOGLE_SERVICE_ACCOUNT_JSON environment variable")
        elif service_account_file and os.path.exists(service_account_file):
            # Load from file
            credentials = service_account.Credentials.from_service_account_file(
                service_account_file, scopes=SCOPES
            )
        else:
            raise ValueError(
                "Service account credentials not found. Please set either:\n"
                "- GOOGLE_SERVICE_ACCOUNT_JSON (JSON string) or\n" 
                "- GOOGLE_SERVICE_ACCOUNT_FILE (file path) environment variable"
            )
        
        # Build the Drive API service
        self.service = build('drive', 'v3', credentials=credentials)
    
    def get_all_files_recursive(self, folder_id: str) -> List[Dict]:
        """
        Recursively get all files from a folder and its subfolders
        
        Args:
            folder_id: The Google Drive folder ID to scan
            
        Returns:
            List of file information dictionaries
        """
        all_files = []
        folders_to_process = [folder_id]
        
        while folders_to_process:
            current_folder = folders_to_process.pop(0)
            
            # Query for all items in current folder
            query = f"'{current_folder}' in parents and trashed=false"
            
            page_token = None
            while True:
                try:
                    results = self.service.files().list(
                        q=query,
                        pageSize=100,
                        fields="nextPageToken, files(id, name, mimeType, size, modifiedTime, webViewLink)",
                        pageToken=page_token
                    ).execute()
                    
                    items = results.get('files', [])
                    
                    for item in items:
                        if item['mimeType'] == 'application/vnd.google-apps.folder':
                            # Add subfolder to processing queue
                            folders_to_process.append(item['id'])
                        else:
                            # Add file to results
                            all_files.append(item)
                    
                    page_token = results.get('nextPageToken')
                    if not page_token:
                        break
                        
                except Exception as e:
                    print(f"Error scanning folder {current_folder}: {str(e)}")
                    break
        
        return all_files
    
    def download_file(self, file_id: str) -> io.BytesIO:
        """
        Download a file from Google Drive
        
        Args:
            file_id: The Google Drive file ID
            
        Returns:
            BytesIO object containing file content
        """
        request = self.service.files().get_media(fileId=file_id)
        file_buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(file_buffer, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        file_buffer.seek(0)
        return file_buffer
    
    def export_google_doc(self, file_id: str, mime_type: str) -> str:
        """
        Export Google Workspace files (Docs, Sheets, Slides) as text
        
        Args:
            file_id: The Google Drive file ID
            mime_type: The Google Workspace MIME type
            
        Returns:
            Text content of the document
        """
        # Map Google Workspace MIME types to export formats
        export_formats = {
            'application/vnd.google-apps.document': 'text/plain',
            'application/vnd.google-apps.spreadsheet': 'text/csv',
            'application/vnd.google-apps.presentation': 'text/plain',
        }
        
        export_mime = export_formats.get(mime_type, 'text/plain')
        
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType=export_mime
            )
            file_buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(file_buffer, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            file_buffer.seek(0)
            return file_buffer.read().decode('utf-8', errors='ignore')
        except Exception as e:
            print(f"Error exporting Google Doc {file_id}: {str(e)}")
            return ""
    
    def extract_content(self, file_info: Dict) -> str:
        """
        Extract text content from various file types
        
        Args:
            file_info: Dictionary containing file metadata
            
        Returns:
            Extracted text content
        """
        file_id = file_info['id']
        mime_type = file_info['mimeType']
        file_name = file_info['name']
        
        try:
            # Handle Google Workspace files
            if mime_type.startswith('application/vnd.google-apps'):
                return self.export_google_doc(file_id, mime_type)
            
            # Download regular files
            file_buffer = self.download_file(file_id)
            
            # Extract based on MIME type
            if mime_type == 'application/pdf':
                return self._extract_pdf(file_buffer)
            
            elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
                return self._extract_docx(file_buffer)
            
            elif mime_type in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]:
                return self._extract_xlsx(file_buffer)
            
            elif mime_type == 'text/csv':
                return self._extract_csv(file_buffer)
            
            elif mime_type.startswith('text/'):
                return file_buffer.read().decode('utf-8', errors='ignore')
            
            else:
                print(f"Unsupported file type: {mime_type} for {file_name}")
                return ""
                
        except Exception as e:
            print(f"Error extracting content from {file_name}: {str(e)}")
            return ""
    
    def _extract_pdf(self, file_buffer: io.BytesIO) -> str:
        """Extract text from PDF file"""
        try:
            pdf_reader = PyPDF2.PdfReader(file_buffer)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            print(f"Error extracting PDF: {str(e)}")
            return ""
    
    def _extract_docx(self, file_buffer: io.BytesIO) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(file_buffer)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            print(f"Error extracting DOCX: {str(e)}")
            return ""
    
    def _extract_xlsx(self, file_buffer: io.BytesIO) -> str:
        """Extract text from XLSX file"""
        try:
            workbook = openpyxl.load_workbook(file_buffer, data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            
            return text
        except Exception as e:
            print(f"Error extracting XLSX: {str(e)}")
            return ""
    
    def _extract_csv(self, file_buffer: io.BytesIO) -> str:
        """Extract text from CSV file"""
        try:
            text = file_buffer.read().decode('utf-8', errors='ignore')
            return text
        except Exception as e:
            print(f"Error extracting CSV: {str(e)}")
            return ""
